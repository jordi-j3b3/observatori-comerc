"""
Processador de dades: obte dades via API, processa i desa a cache/.
Jerarquia de fonts:
  1. API (INE, Eurostat, CNMC) — font primaria per actualitzacions automatiques
  2. CSV cache existent — fallback si l'API falla (mante dades anteriors)
  3. Excel local — nomes per a la carrega inicial (els Excel no es pugen a GitHub)
"""
import os
import pandas as pd
from data.fetchers import ine, eurostat, cnmc

CACHE_DIR = os.path.join(os.path.dirname(__file__), "cache")


def ensure_cache_dir():
    os.makedirs(CACHE_DIR, exist_ok=True)


def save_cache(df, name):
    ensure_cache_dir()
    path = os.path.join(CACHE_DIR, f"{name}.csv")
    df.to_csv(path, index=False)
    print(f"  Desat: {path} ({len(df)} files)")


def load_cache(name):
    path = os.path.join(CACHE_DIR, f"{name}.csv")
    if os.path.exists(path):
        return pd.read_csv(path)
    return pd.DataFrame()


# ─── FONTS EXCEL (CARREGA INICIAL LOCAL) ─────────────────────

def load_excel_pib():
    """Carrega dades del PIB CNAE 47 des de l'Excel existent com a base."""
    try:
        import openpyxl
    except ImportError:
        return pd.DataFrame()

    path = os.path.join(os.path.dirname(__file__), "..", "..", "PIB CNAE 47.xlsx")
    if not os.path.exists(path):
        return pd.DataFrame()

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["PIB "]

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue
        try:
            any_ = int(str(row[0]).strip())
        except (ValueError, TypeError):
            continue

        rec = {"any": any_}
        if row[1] is not None:
            rec["vab_total_corrents"] = float(row[1])
        if row[3] is not None:
            rec["vab_cnae47_corrents"] = float(row[3])
        if len(row) > 9 and row[9] is not None:
            rec["vab_total_constants"] = float(row[9])
        if len(row) > 11 and row[11] is not None:
            rec["vab_cnae47_constants"] = float(row[11])
        if len(row) > 24 and row[24] is not None:
            rec["empreses"] = float(row[24])
        if len(row) > 25 and row[25] is not None:
            rec["pib_per_empresa"] = float(row[25])

        records.append(rec)

    return pd.DataFrame(records)


def load_excel_productivitat():
    """Carrega dades de productivitat des de l'Excel existent."""
    try:
        import openpyxl
    except ImportError:
        return pd.DataFrame()

    path = os.path.join(os.path.dirname(__file__), "..", "..", "PRODUCTIVITAT.xlsx")
    if not os.path.exists(path):
        return pd.DataFrame()

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Hoja1"]

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if len(rows) < 5:
        return pd.DataFrame()

    anys = [int(v) for v in rows[0][1:] if isinstance(v, (int, float))]
    data = {}
    labels = [
        "xifra_negoci_constants", "valor_afegit_constants",
        "personal_ocupat", "hores_treballades"
    ]

    for i, label in enumerate(labels):
        if i + 1 < len(rows):
            vals = [v for v in rows[i + 1][1:] if isinstance(v, (int, float))]
            data[label] = vals[:len(anys)]

    records = []
    for j, any_ in enumerate(anys):
        rec = {"any": any_}
        for label in labels:
            if label in data and j < len(data[label]):
                rec[label] = data[label][j]
        records.append(rec)

    df = pd.DataFrame(records)

    if "valor_afegit_constants" in df.columns and "hores_treballades" in df.columns:
        df["productivitat_va_hora"] = df["valor_afegit_constants"] / df["hores_treballades"]
    if "xifra_negoci_constants" in df.columns and "hores_treballades" in df.columns:
        df["productivitat_xn_hora"] = df["xifra_negoci_constants"] / df["hores_treballades"]

    return df


def load_excel_empreses_cat():
    """Carrega dades d'empreses de Catalunya."""
    try:
        import openpyxl
    except ImportError:
        return pd.DataFrame()

    path = os.path.join(os.path.dirname(__file__), "..", "..", "empreses cnae47 Catalunya.xlsx")
    if not os.path.exists(path):
        return pd.DataFrame()

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["tabla-298"]

    records = []
    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, values_only=True):
        if row[0] and "Cataluny" in str(row[0]):
            for i, val in enumerate(row[1:], start=0):
                if isinstance(val, (int, float)):
                    any_ = 2020 - i
                    records.append({"territori": "catalunya", "any": any_, "empreses": int(val)})
        elif row[0] and "Espany" in str(row[0]):
            for i, val in enumerate(row[1:], start=0):
                if isinstance(val, (int, float)):
                    any_ = 2020 - i
                    records.append({"territori": "espanya", "any": any_, "empreses": int(val)})

    return pd.DataFrame(records)


# ─── FONTS API ────────────────────────────────────────────────

def fetch_ipc_annual():
    """Obte IPC mitjana anual (base 2021=100) des de l'API INE."""
    print("  Obtenint IPC (taula 50902)...")
    try:
        df_ipc = ine.fetch_ipc()
        if df_ipc.empty:
            print("  IPC: sense dades")
            return pd.DataFrame()

        df_ipc["ipc"] = pd.to_numeric(df_ipc["ipc"], errors="coerce")
        df_ipc["any"] = pd.to_numeric(df_ipc["any"], errors="coerce")
        ipc_anual = df_ipc.groupby("any")["ipc"].mean().reset_index()
        ipc_anual.columns = ["any", "ipc_mitjana"]
        print(f"  IPC: {len(ipc_anual)} anys (base 2021=100)")
        return ipc_anual

    except Exception as e:
        print(f"  Error IPC: {e}")
        return pd.DataFrame()


def fetch_pib_vab_from_api():
    """
    Obte PIB/VAB CNAE 47 des de l'API INE (taula 69070).
    Deflacta amb IPC general reancorat al primer any de la serie:
      VAB_constants = VAB_corrents / (IPC_any / IPC_primer_any)
    Així el valor real del primer any = nominal, i després real <= nominal sempre.
    """
    print("  Intentant API INE (taula 69070)...")
    try:
        df_api = ine.fetch_vab_ramas()
        if df_api.empty:
            print("  API INE: sense dades")
            return pd.DataFrame()

        # Pivotar: una fila per any amb columnes per indicador
        pivots = {}
        for ind in df_api["indicador"].unique():
            df_ind = df_api[df_api["indicador"] == ind][["any", "valor"]].copy()
            df_ind = df_ind.rename(columns={"valor": ind})
            pivots[ind] = df_ind

        if "vab_total_corrents" not in pivots or "vab_cnae47_corrents" not in pivots:
            print("  API INE: no s'han trobat series corrents")
            return pd.DataFrame()

        # Merge series corrents
        df = pivots["vab_total_corrents"]
        if "vab_cnae47_corrents" in pivots:
            df = df.merge(pivots["vab_cnae47_corrents"], on="any", how="outer")

        df = df.sort_values("any").reset_index(drop=True)

        # Obtenir IPC per deflactar
        ipc = fetch_ipc_annual()
        if not ipc.empty:
            df = df.merge(ipc, on="any", how="left")

            # Deflactar: reancorar IPC al primer any amb dades
            if "ipc_mitjana" in df.columns:
                ipc_base = df["ipc_mitjana"].dropna().iloc[0]
                df["vab_cnae47_constants"] = df["vab_cnae47_corrents"] / (df["ipc_mitjana"] / ipc_base)
                df["vab_total_constants"] = df["vab_total_corrents"] / (df["ipc_mitjana"] / ipc_base)
                any_base = int(df.loc[df["ipc_mitjana"].first_valid_index(), "any"])
                print(f"  Deflactat amb IPC reancorat a {any_base} (IPC base={ipc_base:.2f})")

            # Netejar columna auxiliar
            df = df.drop(columns=["ipc_mitjana"], errors="ignore")

        print(f"  API INE: {len(df)} anys obtinguts")
        return df

    except Exception as e:
        print(f"  Error API INE: {e}")
        return pd.DataFrame()


def fetch_empreses_from_api():
    """Obte empreses CNAE 47 des de l'API INE (taula 298)."""
    print("  Intentant API INE (taula 298)...")
    try:
        # Espanya + Catalunya
        df_api = ine.fetch_empreses()
        if df_api.empty:
            print("  API INE empreses: sense dades")
            return pd.DataFrame()

        # Agregar per any anual (l'API pot tornar dades amb periode)
        df_api["any"] = pd.to_numeric(df_api["any"], errors="coerce")
        df_api["empreses"] = pd.to_numeric(df_api["empreses"], errors="coerce")
        df = df_api.dropna(subset=["any", "empreses"]).copy()
        df["any"] = df["any"].astype(int)
        df["empreses"] = df["empreses"].astype(int)

        print(f"  API INE empreses: {len(df)} registres")
        return df[["territori", "any", "empreses"]]

    except Exception as e:
        print(f"  Error API INE empreses: {e}")
        return pd.DataFrame()


def fetch_europa_from_api():
    """Obte dades europees des de l'API Eurostat."""
    print("  Intentant API Eurostat...")
    try:
        df_vab47 = eurostat.fetch_vab_cnae47()
        df_vab_total = eurostat.fetch_vab_total()

        if df_vab47.empty or df_vab_total.empty:
            print("  Eurostat: sense dades")
            return pd.DataFrame()

        df = df_vab47.merge(df_vab_total, on=["pais", "pais_codi", "any"], how="inner")
        df["pes_cnae47"] = df["vab_meur"] / df["vab_total_meur"]

        print(f"  Eurostat: {len(df)} registres de {df['pais_codi'].nunique()} paisos")
        return df

    except Exception as e:
        print(f"  Error Eurostat: {e}")
        return pd.DataFrame()


def fetch_ecommerce_from_api():
    """Obte dades d'e-commerce des de l'API CNMC."""
    print("  Intentant API CNMC...")
    try:
        df_api = cnmc.fetch_ecommerce()
        if df_api.empty:
            print("  CNMC: sense dades")
            return pd.DataFrame()
        print(f"  CNMC: {len(df_api)} registres")
        return df_api
    except Exception as e:
        print(f"  Error CNMC: {e}")
        return pd.DataFrame()


def fetch_productivitat_from_api():
    """
    Obte dades de productivitat des de l'API INE.
    Font: Taula 36194 (EEE Sector Comercio) — CNAE 47.
    Metodologia:
      1. Obtenir magnituds a preus corrents (xifra negoci, VA, personal, hores)
      2. Deflactar amb IPC reancorat al primer any: valor_constants = valor_corrents / (IPC/IPC_base)
      3. Calcular productivitat: VA/hora, XN/hora
    """
    print("  Intentant API INE (taula 36194 - EEE Comercio)...")
    try:
        df = ine.fetch_eee_comercio()
        if df.empty:
            print("  API INE EEE: sense dades")
            return pd.DataFrame()

        print(f"  EEE Comercio: {len(df)} anys, magnituds: {list(df.columns)}")

        # Obtenir IPC per deflactar
        ipc = fetch_ipc_annual()
        if ipc.empty:
            print("  No es pot deflactar sense IPC")
            return pd.DataFrame()

        df = df.merge(ipc, on="any", how="left")

        # Deflactar xifra negoci i valor afegit amb IPC reancorat al primer any
        if "ipc_mitjana" in df.columns:
            ipc_base = df["ipc_mitjana"].dropna().iloc[0]
            deflactor = df["ipc_mitjana"] / ipc_base
            if "xifra_negoci" in df.columns:
                df["xifra_negoci_constants"] = df["xifra_negoci"] / deflactor
            if "valor_afegit" in df.columns:
                df["valor_afegit_constants"] = df["valor_afegit"] / deflactor
            any_base = int(df.loc[df["ipc_mitjana"].first_valid_index(), "any"])
            print(f"  Deflactat amb IPC reancorat a {any_base}")

        # Calcular productivitat
        if "valor_afegit_constants" in df.columns and "hores_treballades" in df.columns:
            df["productivitat_va_hora"] = df["valor_afegit_constants"] / df["hores_treballades"]
        if "xifra_negoci_constants" in df.columns and "hores_treballades" in df.columns:
            df["productivitat_xn_hora"] = df["xifra_negoci_constants"] / df["hores_treballades"]

        # ── Afegir dades de Resultats d'explotacio (T=36199) ──
        try:
            df_pyl = ine.fetch_eee_pyl()
            if not df_pyl.empty:
                # Merge per any (mantenim columnes existents si hi ha conflicte)
                cols_pyl = [c for c in ["cogs", "serveis_exteriors", "sueldos_salaris"]
                            if c in df_pyl.columns]
                df = df.merge(df_pyl[["any"] + cols_pyl], on="any", how="left")
                print(f"  Resultats d'explotacio (T=36199): {len(df_pyl)} anys integrats")
        except Exception as e:
            print(f"  Error obtenint T=36199: {e}")

        # Deflactar gastos de personal
        if "gastos_personal" in df.columns and "ipc_mitjana" in df.columns:
            df["gastos_personal_constants"] = df["gastos_personal"] / deflactor

        # Deflactar COGS i serveis exteriors per fer ratios consistents en preus constants
        if "cogs" in df.columns and "ipc_mitjana" in df.columns:
            df["cogs_constants"] = df["cogs"] / deflactor
        if "serveis_exteriors" in df.columns and "ipc_mitjana" in df.columns:
            df["serveis_exteriors_constants"] = df["serveis_exteriors"] / deflactor

        # Marge brut comptable: (Vendes - COGS) / Vendes (a preus corrents = ratio invariant)
        if "cogs" in df.columns and "xifra_negoci" in df.columns:
            df["marge_brut"] = (df["xifra_negoci"] - df["cogs"]) / df["xifra_negoci"]

        # Distribucio del VAB: quota salarial i excedent brut
        if "gastos_personal" in df.columns and "valor_afegit" in df.columns:
            df["quota_salarial"] = df["gastos_personal"] / df["valor_afegit"]
            df["excedent_brut"] = df["valor_afegit"] - df["gastos_personal"]
        if "gastos_personal" in df.columns and "personal_ocupat" in df.columns:
            df["cost_laboral_per_ocupat"] = df["gastos_personal"] / df["personal_ocupat"]
        if "gastos_personal_constants" in df.columns and "hores_treballades" in df.columns:
            df["cost_laboral_hora"] = df["gastos_personal_constants"] / df["hores_treballades"]

        # Seleccionar columnes finals
        cols_finals = ["any", "xifra_negoci_constants", "valor_afegit_constants",
                       "personal_ocupat", "hores_treballades",
                       "productivitat_va_hora", "productivitat_xn_hora",
                       "gastos_personal", "gastos_personal_constants",
                       "quota_salarial", "excedent_brut",
                       "cost_laboral_per_ocupat", "cost_laboral_hora",
                       "cogs", "cogs_constants", "serveis_exteriors",
                       "serveis_exteriors_constants", "marge_brut"]
        cols_disponibles = [c for c in cols_finals if c in df.columns]
        df = df[cols_disponibles]

        print(f"  Productivitat calculada: {len(df)} anys")
        return df

    except Exception as e:
        print(f"  Error API INE productivitat: {e}")
        return pd.DataFrame()


# ─── PROCESSAMENT (API → CACHE → EXCEL) ──────────────────────

def process_pib_vab():
    """
    Processa PIB/VAB.
    1. Intenta API INE
    2. Fallback: cache existent
    3. Fallback: Excel local
    """
    print("  Font 1: API INE")
    df = fetch_pib_vab_from_api()

    if df.empty:
        print("  Font 2: Cache existent")
        df = load_cache("pib_vab")

    if df.empty:
        print("  Font 3: Excel local")
        df = load_excel_pib()

    if df.empty:
        print("  Cap font disponible per PIB/VAB")
        return pd.DataFrame()

    # Calcular metriques derivades
    if "vab_cnae47_corrents" in df.columns and "vab_total_corrents" in df.columns:
        df["pes_cnae47"] = df["vab_cnae47_corrents"] / df["vab_total_corrents"]

    for col in ["vab_cnae47_corrents", "vab_cnae47_constants"]:
        if col in df.columns:
            df[f"var_{col}"] = df[col].pct_change()

    save_cache(df, "pib_vab")
    return df


def process_empreses():
    """
    Processa empreses.
    1. Intenta API INE (Espanya + Catalunya)
    2. Fallback: cache existent
    3. Fallback: Excel local
    """
    print("  Font 1: API INE")
    df = fetch_empreses_from_api()

    if df.empty:
        print("  Font 2: Cache existent")
        df = load_cache("empreses")

    if df.empty:
        print("  Font 3: Excel local")
        df_pib = load_excel_pib()
        df_cat = load_excel_empreses_cat()

        records = []
        if not df_pib.empty and "empreses" in df_pib.columns:
            for _, row in df_pib.iterrows():
                if pd.notna(row.get("empreses")):
                    records.append({
                        "territori": "espanya",
                        "any": int(row["any"]),
                        "empreses": int(row["empreses"]),
                    })

        if not df_cat.empty:
            for _, row in df_cat.iterrows():
                records.append({
                    "territori": row["territori"],
                    "any": int(row["any"]),
                    "empreses": int(row["empreses"]),
                })

        df = pd.DataFrame(records)

    if df.empty:
        print("  Cap font disponible per empreses")
        return pd.DataFrame()

    # ── Afegir població i densitat comercial ──
    print("  Obtenint dades de població...")
    try:
        df_pob = ine.fetch_poblacio()
        if not df_pob.empty:
            # Homogeneitzar noms de territori (Padrón vs DIRCE)
            NORM = {
                "Asturias, Principado de": "Asturias (Principado de)",
                "Balears, Illes": "Balears (Illes)",
                "Madrid, Comunidad de": "Madrid (Comunidad de)",
                "Murcia, Región de": "Murcia (Región de)",
                "Navarra, Comunidad Foral de": "Navarra (Comunidad Foral de)",
                "Rioja, La": "Rioja (La)",
            }
            df_pob["territori"] = df_pob["territori"].replace(NORM)
            df_pob["any"] = pd.to_numeric(df_pob["any"], errors="coerce")
            df_pob["poblacio"] = pd.to_numeric(df_pob["poblacio"], errors="coerce")

            df["any"] = pd.to_numeric(df["any"], errors="coerce")
            df = df.merge(df_pob[["territori", "any", "poblacio"]], on=["territori", "any"], how="left")

            # Empreses per 1.000 habitants
            mask = df["poblacio"].notna() & (df["poblacio"] > 0)
            df.loc[mask, "empreses_per_1000hab"] = (df.loc[mask, "empreses"] / df.loc[mask, "poblacio"]) * 1000

            n_with = df["empreses_per_1000hab"].notna().sum()
            print(f"  Densitat comercial calculada: {n_with} registres")
    except Exception as e:
        print(f"  Error obtenint població: {e}")

    save_cache(df, "empreses")
    return df


def process_productivitat():
    """
    Processa productivitat.
    Nota: les dades de productivitat de l'INE requereixen transformacio
    complexa. Prioritzem cache > Excel > API.
    """
    print("  Font 1: API INE")
    df = fetch_productivitat_from_api()

    if df.empty:
        print("  Font 2: Cache existent")
        df = load_cache("productivitat")

    if df.empty:
        print("  Font 3: Excel local")
        df = load_excel_productivitat()

    if df.empty:
        print("  Cap font disponible per productivitat")
        return pd.DataFrame()

    save_cache(df, "productivitat")
    return df


def process_ecommerce():
    """
    Processa e-commerce.
    1. Intenta API CNMC
    2. Fallback: cache existent
    3. Fallback: Excel local
    """
    print("  Font 1: API CNMC")
    df = fetch_ecommerce_from_api()

    if df.empty:
        print("  Font 2: Cache existent")
        df = load_cache("ecommerce")

    if df.empty:
        print("  Font 3: Excel local")
        df = cnmc.fetch_ecommerce_from_excel()

    if df.empty:
        print("  Cap font disponible per e-commerce")
        return pd.DataFrame()

    if "ecommerce_total_eur" in df.columns and "ecommerce_cnae47_eur" in df.columns:
        df["pes_cnae47_ecommerce"] = df["ecommerce_cnae47_eur"] / df["ecommerce_total_eur"]

    save_cache(df, "ecommerce")
    return df


def process_europa():
    """
    Processa comparativa europea.
    1. Intenta API Eurostat
    2. Fallback: cache existent
    3. Fallback: Excel local
    """
    COUNTRY_NAMES_CA = {
        "EU27_2020": "UE-27", "EA20": "Eurozona",
        "BE": "Belgica", "BG": "Bulgaria", "CZ": "Txequia",
        "DK": "Dinamarca", "DE": "Alemanya", "EE": "Estonia", "IE": "Irlanda",
        "EL": "Grecia", "ES": "Espanya", "FR": "Franca", "HR": "Croacia",
        "IT": "Italia", "CY": "Xipre", "LV": "Letonia", "LT": "Lituania",
        "LU": "Luxemburg", "HU": "Hongria", "MT": "Malta", "NL": "Paisos Baixos",
        "AT": "Austria", "PL": "Polonia", "PT": "Portugal", "RO": "Romania",
        "SI": "Eslovenia", "SK": "Eslovaquia", "FI": "Finlandia", "SE": "Suecia",
        "IS": "Islandia", "NO": "Noruega", "CH": "Suissa", "RS": "Serbia",
        "BA": "Bosnia",
    }

    print("  Font 1: API Eurostat")
    df = fetch_europa_from_api()

    if not df.empty:
        # Assegurar noms en catala
        df["pais"] = df["pais_codi"].map(COUNTRY_NAMES_CA).fillna(df["pais"])

    if df.empty:
        print("  Font 2: Cache existent")
        df = load_cache("europa_vab")

    if df.empty:
        print("  Font 3: Excel local")
        df = _load_europa_from_excel(COUNTRY_NAMES_CA)

    if df.empty:
        print("  Cap font disponible per Europa")
        return pd.DataFrame()

    save_cache(df, "europa_vab")
    return df


def _load_europa_from_excel(country_names_ca):
    """Llegeix dades europees de l'Excel local (nomes primera carrega)."""
    try:
        import openpyxl
    except ImportError:
        return pd.DataFrame()

    path = os.path.join(os.path.dirname(__file__), "..", "..", "nama_10_a64__custom_16418785_spreadsheet.xlsx")
    if not os.path.exists(path):
        return pd.DataFrame()

    wb = openpyxl.load_workbook(path, data_only=True)
    if "PIB UE CNAE47" not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb["PIB UE CNAE47"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    if not rows:
        return pd.DataFrame()

    header = rows[0]
    anys = []
    for v in header[1:]:
        if v is None:
            continue
        try:
            if isinstance(v, str) and "Variaci" in v:
                break
            anys.append(int(v))
        except (ValueError, TypeError):
            continue

    COUNTRY_CODES = {
        "European Union - 27 countries (from 2020)": "EU27_2020",
        "EU - 27 countries": "EU27_2020",
        "Belgium": "BE", "Bulgaria": "BG", "Czechia": "CZ", "Denmark": "DK",
        "Germany": "DE", "Estonia": "EE", "Ireland": "IE", "Greece": "EL",
        "Spain": "ES", "France": "FR", "Croatia": "HR", "Italy": "IT",
        "Cyprus": "CY", "Latvia": "LV", "Lithuania": "LT", "Luxembourg": "LU",
        "Hungary": "HU", "Malta": "MT", "Netherlands": "NL", "Austria": "AT",
        "Poland": "PL", "Portugal": "PT", "Romania": "RO", "Slovenia": "SI",
        "Slovakia": "SK", "Finland": "FI", "Sweden": "SE",
        "Iceland": "IS", "Norway": "NO", "Switzerland": "CH",
        "Serbia": "RS", "Bosnia and Herzegovina": "BA",
    }

    records = []
    seen_countries = set()
    for row in rows[1:]:
        pais_name = row[0]
        if pais_name is None:
            continue
        pais_name = str(pais_name).strip()
        pais_codi = COUNTRY_CODES.get(pais_name)
        if not pais_codi:
            continue

        if pais_codi in seen_countries:
            continue
        seen_countries.add(pais_codi)

        for i, any_ in enumerate(anys):
            idx = i + 1
            if idx < len(row) and row[idx] is not None:
                try:
                    pes = float(row[idx])
                    if pes > 1:
                        continue
                    records.append({
                        "pais": country_names_ca.get(pais_codi, pais_name),
                        "pais_codi": pais_codi,
                        "any": any_,
                        "pes_cnae47": pes,
                    })
                except (ValueError, TypeError):
                    continue

    return pd.DataFrame(records)


def _estimate_vab_eurostat(df):
    """
    Estimació híbrida del VAB CNAE 47 per CCAA combinant:
    - Eurostat nama_10r_3gva: VAB secció G-I per CCAA (comptabilitat regional real)
    - Eurostat nama_10_a64: VAB G47 nacional (per calcular ratio G47/GI)
    - INE EEE taula 76817: xifra de negoci per CCAA (per ponderar)

    Metode:
    1. ratio_g47_gi = VAB_G47_ES / VAB_GI_ES (pes del detall dins G-I a nivell nacional)
    2. Base: VAB_G47_ccaa = VAB_GI_ccaa * ratio_g47_gi
    3. Ajust proporcional amb quotes de xifra de negoci per reflectir que
       el pes del retail dins G-I varia per CCAA (ex: turisme a Balears pesa més en G-I)
    4. Restriccio: la suma de VAB estimat per CCAA = VAB G47 nacional d'Eurostat
    """
    print("  Estimant VAB via Eurostat (metode hibrid)...")
    try:
        df_gi = eurostat.fetch_vab_regional_gi()
        df_g47 = eurostat.fetch_vab_nacional_g47()
        df_total = eurostat.fetch_vab_regional_total()
    except Exception as e:
        print(f"  Error Eurostat regional: {e}")
        return df

    if df_gi.empty or df_g47.empty:
        print("  Eurostat regional: sense dades suficients")
        return df

    gi_es = df_gi[df_gi["territori"] == "espanya"][["any", "vab_gi_meur"]].rename(
        columns={"vab_gi_meur": "vab_gi_es"})
    gi_ccaa = df_gi[df_gi["territori"] != "espanya"][["territori", "any", "vab_gi_meur"]]

    ratio = gi_es.merge(df_g47, on="any", how="inner")
    ratio["r_g47_gi"] = ratio["vab_g47_meur"] / ratio["vab_gi_es"]
    print(f"  Ratio G47/GI nacional: {dict(zip(ratio['any'].astype(int), ratio['r_g47_gi'].round(4)))}")

    df = df.merge(gi_ccaa, on=["territori", "any"], how="left")
    df = df.merge(ratio[["any", "r_g47_gi", "vab_g47_meur"]], on="any", how="left")

    # Base estimate: proportional to G-I
    df["_vab_base"] = df["vab_gi_meur"] * df["r_g47_gi"]

    # XN-weighted adjustment: redistribute so CCAA with higher retail XN share
    # get proportionally more. This corrects for tourism-heavy CCAA where G-I
    # is dominated by hospitality rather than retail.
    df_nac = df[df["territori"] == "espanya"]
    xn_nac_by_year = df_nac[["any", "xifra_negoci"]].rename(
        columns={"xifra_negoci": "xn_nac"}).dropna()

    df = df.merge(xn_nac_by_year, on="any", how="left")

    mask_calc = (
        df["territori"] != "espanya") & (
        df["vab_gi_meur"].notna()) & (
        df["r_g47_gi"].notna()) & (
        df["xifra_negoci"].notna()) & (
        df["xn_nac"].notna()
    )

    for year in df.loc[mask_calc, "any"].unique():
        ym = mask_calc & (df["any"] == year)
        if ym.sum() == 0:
            continue

        # G-I shares and XN shares per CCAA
        gi_total = df.loc[ym, "vab_gi_meur"].sum()
        gi_shares = df.loc[ym, "vab_gi_meur"] / gi_total
        xn_total = df.loc[ym, "xifra_negoci"].sum()
        xn_shares = df.loc[ym, "xifra_negoci"] / xn_total

        # Hybrid share: average of G-I share (top-down) and XN share (bottom-up)
        hybrid_shares = (gi_shares + xn_shares) / 2
        hybrid_shares = hybrid_shares / hybrid_shares.sum()

        # Distribute the national G47 total using hybrid shares
        vab_g47_es = df.loc[ym, "vab_g47_meur"].iloc[0]
        df.loc[ym, "vab_eurostat_meur"] = hybrid_shares * vab_g47_es

    # Also set national total
    mask_es = (df["territori"] == "espanya") & df["vab_g47_meur"].notna()
    df.loc[mask_es, "vab_eurostat_meur"] = df.loc[mask_es, "vab_g47_meur"]

    # Convert M EUR to EUR for consistency with other columns
    if "vab_eurostat_meur" in df.columns:
        df["vab_eurostat"] = df["vab_eurostat_meur"] * 1e6
        n = df["vab_eurostat"].notna().sum()
        print(f"  VAB Eurostat estimat per {n} registres")

    # Pes del CNAE 47 sobre el PIB de cada CCAA
    if not df_total.empty:
        df = df.merge(
            df_total[["territori", "any", "vab_total_meur"]],
            on=["territori", "any"], how="left")
        mask_pes = df["vab_eurostat_meur"].notna() & df["vab_total_meur"].notna() & (df["vab_total_meur"] > 0)
        df.loc[mask_pes, "pes_cnae47_pib"] = df.loc[mask_pes, "vab_eurostat_meur"] / df.loc[mask_pes, "vab_total_meur"]
        n_pes = df["pes_cnae47_pib"].notna().sum()
        print(f"  Pes CNAE 47 / PIB calculat per {n_pes} registres")
    else:
        print("  VAB total regional no disponible, pes no calculat")

    df = df.drop(columns=["vab_gi_meur", "r_g47_gi", "vab_g47_meur", "_vab_base",
                           "xn_nac", "vab_eurostat_meur", "vab_total_meur"], errors="ignore")

    return df


def process_eee_ccaa():
    """
    Processa dades EEE Comercio per CCAA (taula 76817).
    Tres estimacions de VAB CNAE 47 per CCAA:
    - vab_estimat: via ratio nacional VA/XN (EEE, preus constants)
    - vab_estimat_nominal: via ratio nacional VAB_corrents/XN (INE CNA + EEE)
    - vab_eurostat: metode hibrid Eurostat (comptabilitat regional G-I + EEE XN)
    """
    print("  Font 1: API INE (taula 76817)")
    try:
        df = ine.fetch_eee_comercio_ccaa()
        if df.empty:
            print("  API INE EEE CCAA: sense dades")
            df = load_cache("eee_ccaa")
    except Exception as e:
        print(f"  Error API INE EEE CCAA: {e}")
        df = load_cache("eee_ccaa")

    if df.empty:
        print("  Cap font disponible per EEE CCAA")
        return pd.DataFrame()

    print(f"  EEE CCAA: {len(df)} registres, {df['territori'].nunique()} territoris")

    df_nac = df[df["territori"] == "espanya"].copy()
    df_prod = load_cache("productivitat")
    df_pib = load_cache("pib_vab")

    if not df_nac.empty and "xifra_negoci" in df_nac.columns:
        xn_nac = df_nac[["any", "xifra_negoci"]].rename(columns={"xifra_negoci": "xn_nac"})

        # VAB real (preus constants) via productivitat
        if not df_prod.empty and "valor_afegit_constants" in df_prod.columns:
            ratio_real = df_prod[["any"]].copy()
            ratio_real = ratio_real.merge(xn_nac, on="any", how="inner")
            ratio_real["va_real"] = df_prod.set_index("any").loc[ratio_real["any"].values, "valor_afegit_constants"].values
            ratio_real["ratio"] = ratio_real["va_real"] / ratio_real["xn_nac"]

            df = df.merge(ratio_real[["any", "ratio"]].rename(columns={"ratio": "_r_real"}), on="any", how="left")
            mask = df["_r_real"].notna() & df["xifra_negoci"].notna()
            df.loc[mask, "vab_estimat"] = df.loc[mask, "xifra_negoci"] * df.loc[mask, "_r_real"]
            df = df.drop(columns=["_r_real"])
            print(f"  VAB real estimat per {df['vab_estimat'].notna().sum()} registres")

        # VAB nominal (preus corrents) via pib_vab
        if not df_pib.empty and "vab_cnae47_corrents" in df_pib.columns:
            ratio_nom = df_pib[["any", "vab_cnae47_corrents"]].copy()
            ratio_nom = ratio_nom.merge(xn_nac, on="any", how="inner")
            ratio_nom["ratio"] = ratio_nom["vab_cnae47_corrents"] * 1e6 / ratio_nom["xn_nac"]

            df = df.merge(ratio_nom[["any", "ratio"]].rename(columns={"ratio": "_r_nom"}), on="any", how="left")
            mask = df["_r_nom"].notna() & df["xifra_negoci"].notna()
            df.loc[mask, "vab_estimat_nominal"] = df.loc[mask, "xifra_negoci"] * df.loc[mask, "_r_nom"]
            df = df.drop(columns=["_r_nom"])
            print(f"  VAB nominal estimat per {df['vab_estimat_nominal'].notna().sum()} registres")

    # Estimació híbrida Eurostat (comptabilitat regional + EEE)
    df = _estimate_vab_eurostat(df)

    save_cache(df, "eee_ccaa")
    return df


def process_municipal():
    """
    Combina dades municipals de 3 fonts INE per construir un index relatiu de
    capacitat comercial municipal:
      - DIRCE municipal T=4721 (empreses Total + sector G-I)
      - Atlas Renda T=30896 (renda neta mitjana per llar)
      - Padron T=33167 (poblacio per municipi)

    AVIS: l'index NO es VAB CNAE 47 estricte (impossible amb dades publiques).
    Es una proxy d'activitat comercial+associada per municipi (G-I inclou
    transport i hostaleria).

    Genera CSV: municipal.csv amb columnes:
      municipi, any_dirce, total_empreses, empreses_g_i, any_renda,
      renda_llar, any_pob, poblacio, gi_per_1000hab, despesa_potencial,
      index_capacitat (0-100, normalitzat)
    """
    print("  Font 1: DIRCE municipal (T=4721)")
    try:
        df_emp = ine.fetch_empreses_municipal()
        df_emp = df_emp.rename(columns={"any": "any_dirce"})
        print(f"  Empreses: {len(df_emp)} municipis")
    except Exception as e:
        print(f"  Error: {e}")
        df_emp = pd.DataFrame()

    # NOTA: Atlas Renda esta fragmentat en ~52 taules INE (1 per provincia).
    # Saltat al MVP per evitar fer 52 crides API. Pendent per a futura iteracio.
    df_renda = pd.DataFrame()

    print("  Font 2: Padron (T=33167)")
    try:
        df_pob = ine.fetch_poblacio_municipal()
        print(f"  Poblacio: {len(df_pob)} municipis")
    except Exception as e:
        print(f"  Error: {e}")
        df_pob = pd.DataFrame()

    if df_emp.empty:
        print("  Sense dades d'empreses, abortant")
        return pd.DataFrame()

    # Merge per nom de municipi (basic match - alguns no coincidiran per accents/format)
    df = df_emp.copy()
    if not df_renda.empty:
        df = df.merge(df_renda, on="municipi", how="left")
    if not df_pob.empty:
        df = df.merge(df_pob, on="municipi", how="left")

    # Indicadors derivats
    if "empreses_g_i" in df.columns and "poblacio" in df.columns:
        mask = df["poblacio"].notna() & (df["poblacio"] > 0)
        df.loc[mask, "gi_per_1000hab"] = df.loc[mask, "empreses_g_i"] / df.loc[mask, "poblacio"] * 1000

    # Index de capacitat comercial: combinacio normalitzada de:
    #   - empreses G-I absolutes (escala) — log scale
    #   - densitat (qualitat per habitant) — escala lineal
    # Pesos: 50% escala + 50% densitat
    # Filtre: nomes municipis amb poblacio >= 5.000 hab. (eviten falsos positius
    # per duplicats de noms com "Sada", "Mieres", "Cieza", etc.)
    if "empreses_g_i" in df.columns and "gi_per_1000hab" in df.columns:
        mask = (df["empreses_g_i"].notna() & df["gi_per_1000hab"].notna()
                & (df["empreses_g_i"] > 0) & (df["poblacio"] >= 5000))
        sub = df.loc[mask].copy()
        if not sub.empty:
            import numpy as np
            log_emp = np.log10(sub["empreses_g_i"])
            dens = sub["gi_per_1000hab"]
            # Normalitzar 0-100
            def norm(x):
                rng = x.max() - x.min()
                return (x - x.min()) / rng * 100 if rng > 0 else x * 0
            score = (norm(log_emp) + norm(dens)) / 2
            df.loc[mask, "index_capacitat"] = score.values

    save_cache(df, "municipal")
    print(f"  Municipal: {len(df)} files, {df['index_capacitat'].notna().sum() if 'index_capacitat' in df.columns else 0} amb index complet")
    return df


def process_subsectors():
    """
    Processa subsectors CNAE 47:
      - DIRCE (T=73019): empreses per subsector 471-479
      - EAS (T=76818): xifra negoci, VA, personal, inversio per subsector
      - EPF (T=75003): despesa familiar per categoria COICOP

    Genera tres CSVs separats al cache (subsectors_dirce, subsectors_eas, subsectors_epf).
    Cada font te schemas diferents - es deixa la combinacio a la pagina de visualitzacio.
    """
    # ── DIRCE ──
    print("  Font 1: DIRCE (T=73019)")
    try:
        df_dirce = ine.fetch_dirce_subsectors()
        if not df_dirce.empty:
            df_dirce["nom"] = df_dirce["codi"].map(ine.CNAE_47_SUBSECTORS).fillna("CNAE 47 total")
            save_cache(df_dirce, "subsectors_dirce")
            print(f"  DIRCE: {len(df_dirce)} files, anys {df_dirce['any'].min()}-{df_dirce['any'].max()}")
        else:
            print("  DIRCE: sense dades, mantenint cache")
    except Exception as e:
        print(f"  Error DIRCE: {e}")

    # ── EAS ──
    print("  Font 2: EAS / EEE Comercio (T=76818)")
    try:
        df_eas = ine.fetch_eas_subsectors()
        if not df_eas.empty:
            df_eas["nom"] = df_eas["codi"].map(ine.CNAE_47_SUBSECTORS).fillna("CNAE 47 total")
            save_cache(df_eas, "subsectors_eas")
            print(f"  EAS: {len(df_eas)} files, anys {df_eas['any'].min()}-{df_eas['any'].max()}")
        else:
            print("  EAS: sense dades, mantenint cache")
    except Exception as e:
        print(f"  Error EAS: {e}")

    # ── EPF ──
    print("  Font 3: EPF (T=75003)")
    try:
        df_epf = ine.fetch_epf_coicop()
        if not df_epf.empty:
            df_epf["nom"] = df_epf["codi_coicop"].map(ine.COICOP_GROUPS).fillna("Total despesa")
            save_cache(df_epf, "subsectors_epf")
            print(f"  EPF: {len(df_epf)} files, anys {df_epf['any'].min()}-{df_epf['any'].max()}")
        else:
            print("  EPF: sense dades, mantenint cache")
    except Exception as e:
        print(f"  Error EPF: {e}")


def save_last_update():
    """Desa la data i hora de l'última actualització."""
    from datetime import datetime
    ensure_cache_dir()
    path = os.path.join(CACHE_DIR, "last_update.txt")
    with open(path, "w") as f:
        f.write(datetime.now().strftime("%Y-%m-%d %H:%M"))
    print(f"  Data actualització: {datetime.now().strftime('%d/%m/%Y %H:%M')}")


def process_all():
    """Executa tot el processament."""
    print("Processant dades...")
    print("Jerarquia: API -> Cache CSV -> Excel local\n")

    print("1. PIB i VAB:")
    process_pib_vab()

    print("\n2. Empreses:")
    process_empreses()

    print("\n3. Productivitat:")
    process_productivitat()

    print("\n4. E-commerce:")
    process_ecommerce()

    print("\n5. Europa:")
    process_europa()

    print("\n6. EEE Comercio per CCAA:")
    process_eee_ccaa()

    print("\n7. Subsectors CNAE 47 (DIRCE + EAS + EPF):")
    process_subsectors()

    print("\n8. Registrant data actualitzacio:")
    save_last_update()

    print("\nProcessament complet!")


if __name__ == "__main__":
    process_all()
