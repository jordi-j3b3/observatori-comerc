"""
Client API per la CNMC (Comision Nacional de los Mercados y la Competencia).
Dades d'e-commerce. Descarga directa CSV.
"""
import requests
import pandas as pd
import io

# ID del recurs CSV: volum de negoci del comerc electronic per rames d'activitat
# Actualitzat 2026-07-06: l'ID anterior (ec53a970-07c9-4dac-8fb4-bdabfd7f2808,
# dataset ds_14116_1) retornava 404. La CNMC ha tornat a reidentificar el recurs
# dins el dataset ds_24264_1 ("Evolución trimestral del volumen de negocio del
# comercio electrónico por ramas de actividad"), mateixes columnes i format.
# Historial: fddca89f… (403) → ec53a970… (404) → recurs actual.
VOLUM_NEGOCI_CSV_ID = "7b7e963f-a8d8-4897-bd19-d96a086a6019"
DUMP_URL = "https://catalogodatos.cnmc.es/datastore/dump"

# Codis CNAE que pertanyen al comerç minorista (47)
CNAE_47_CODES = {
    "4711", "4712", "4724", "4725", "4726", "4727",
    "4730", "4740",
    "4753 + 4755", "4754",
    "4761 + 4769", "4763 + 4764",
    "4771", "4772", "4773", "4774", "4775", "4776", "4777", "4778", "4779",
    "4782", "4791", "4792",
    # Codis mixtos que inclouen 47
    "4646 + 4773", "4647", "4648 + 4777",
    "4684 + 4752",
    "470069", "470083",
}


def fetch_ecommerce():
    """
    Descarrega dades trimestrals d'e-commerce de la CNMC.
    Retorna volum de negoci anual: total i CNAE 47.
    """
    url = f"{DUMP_URL}/{VOLUM_NEGOCI_CSV_ID}"
    # CNMC bloqueja peticions sense User-Agent (retorna 403/500 al UA per defecte
    # de python-requests). Forcem un UA de navegador.
    headers = {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 14.0) "
                      "AppleWebKit/605.1.15 (KHTML, like Gecko) "
                      "Observatori-Comerc/1.0",
        "Accept": "text/csv,application/json,*/*",
    }
    resp = requests.get(url, headers=headers, timeout=120)
    if resp.status_code != 200:
        print(f"CNMC error {resp.status_code}")
        return pd.DataFrame()

    df = pd.read_csv(io.StringIO(resp.text))

    if df.empty:
        return pd.DataFrame()

    col_volum = "evolucion_trimestral_del_volumen_de_negocio"
    col_code = "codigo_armonizacion_segun_cnae_y_cnpa_"
    col_trim = "trimestre"

    if col_volum not in df.columns:
        print(f"CNMC: columna '{col_volum}' no trobada")
        return pd.DataFrame()

    # Extreure any del trimestre (format "2013T4")
    df["any"] = df[col_trim].str[:4].astype(int)
    df[col_volum] = pd.to_numeric(df[col_volum], errors="coerce")

    # Total per any (totes les rames)
    df_total = df.groupby("any")[col_volum].sum().reset_index()
    df_total.columns = ["any", "ecommerce_total_eur"]

    # CNAE 47 per any
    df_47 = df[df[col_code].isin(CNAE_47_CODES)]
    df_47_annual = df_47.groupby("any")[col_volum].sum().reset_index()
    df_47_annual.columns = ["any", "ecommerce_cnae47_eur"]

    # Merge
    result = df_total.merge(df_47_annual, on="any", how="left")
    result = result.sort_values("any")

    return result


def fetch_ecommerce_from_excel():
    """
    Alternativa: llegir dades d'e-commerce des dels Excel existents
    (per quan l'API CNMC no funciona o les dades no estan al datastore).
    """
    import os
    try:
        import openpyxl
    except ImportError:
        return pd.DataFrame()

    excel_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "..",
        "PIB CNAE 47.xlsx"
    )

    if not os.path.exists(excel_path):
        return pd.DataFrame()

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if "Dades ecommerce" not in wb.sheetnames:
        return pd.DataFrame()

    ws = wb["Dades ecommerce"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    if len(rows) < 4:
        return pd.DataFrame()

    anys = [v for v in rows[1] if isinstance(v, (int, float))]
    total = [v for v in rows[2] if isinstance(v, (int, float))]
    cnae47 = [v for v in rows[3] if isinstance(v, (int, float))]

    records = []
    for i, any_ in enumerate(anys):
        rec = {"any": int(any_)}
        if i < len(total):
            rec["ecommerce_total_eur"] = total[i]
        if i < len(cnae47):
            rec["ecommerce_cnae47_eur"] = cnae47[i]
        records.append(rec)

    return pd.DataFrame(records)


if __name__ == "__main__":
    print("Testejant CNMC...")
    print("\nDescarregant volum de negoci e-commerce:")
    df = fetch_ecommerce()
    print(f"   {len(df)} anys")
    if not df.empty:
        print(df)
